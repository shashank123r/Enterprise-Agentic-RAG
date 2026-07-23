"""XLSX extractor — sync calls offloaded via run_in_executor."""

from app.core.logging import get_logger
from app.ingestion.executor import run_in_executor
from app.ingestion.extractors import (
    BaseExtractor, ExtractionResult, PageResult, TableResult, extractor_registry,
)

logger = get_logger(__name__)


@extractor_registry.register("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
class XlsxExtractor(BaseExtractor):
    """Extract data from XLSX spreadsheets."""

    async def extract(self) -> ExtractionResult:
        def _do() -> ExtractionResult:
            from openpyxl import load_workbook
            wb = load_workbook(str(self.file_path), read_only=True, data_only=True)
            r = ExtractionResult()
            props = wb.properties
            r.metadata.update({"title": props.title or "", "author": props.creator or "", "created": str(props.created) if props.created else "", "modified": str(props.modified) if props.modified else "", "sheet_count": len(wb.sheetnames), "sheet_names": list(wb.sheetnames)})
            table_idx = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_data = [[str(c) if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
                if not rows_data: continue
                headers = rows_data[0]
                data_rows = rows_data[1:] if len(rows_data) > 1 else []
                preview = f"=== Sheet: {sheet_name} ===\n" + " | ".join(headers) + "\n" + "\n".join(" | ".join(row) for row in data_rows[:50])
                r.pages.append(PageResult(page_number=table_idx + 1, text=preview, section_title=sheet_name))
                r.tables.append(TableResult(page_number=table_idx + 1, table_index=table_idx, headers=headers, rows=data_rows, caption=f"Sheet: {sheet_name}"))
                table_idx += 1
            r.text = "\n\n".join(p.text for p in r.pages)
            return r
        return await run_in_executor(_do)

    async def extract_text(self) -> str:
        def _do() -> str:
            from openpyxl import load_workbook
            wb = load_workbook(str(self.file_path), read_only=True, data_only=True)
            texts: list[str] = []
            for sheet_name in wb.sheetnames:
                texts.append(f"=== {sheet_name} ===")
                for row in wb[sheet_name].iter_rows(values_only=True):
                    line = " | ".join(str(c) for c in row if c is not None)
                    if line.strip(): texts.append(line)
            return "\n".join(texts)
        return await run_in_executor(_do)

    async def extract_metadata(self) -> dict:
        def _do() -> dict:
            from openpyxl import load_workbook
            wb = load_workbook(str(self.file_path), read_only=True, data_only=True)
            props = wb.properties
            return {"title": props.title or "", "author": props.creator or "", "created": str(props.created) if props.created else "", "modified": str(props.modified) if props.modified else "", "sheet_count": len(wb.sheetnames), "sheet_names": list(wb.sheetnames)}
        return await run_in_executor(_do)
